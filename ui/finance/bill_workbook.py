from io import BytesIO
from zoneinfo import ZoneInfo

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TOTAL_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
AMOUNT_DUE_HEADER_FILL = PatternFill("solid", fgColor="C00000")
AMOUNT_DUE_FILL = PatternFill("solid", fgColor="FCE8E6")
AMOUNT_DUE_TOTAL_FILL = PatternFill("solid", fgColor="F4CCCC")
AMOUNT_DUE_FONT = Font(color="C00000", bold=True)
NEW_YORK = ZoneInfo("America/New_York")
MINIMUM_COLUMN_WIDTHS = {
    "商品 / 材质": 24,
    "型号": 18,
    "订单数": 12,
    "商品数量": 14,
    "数量": 12,
    "正确单价": 18,
    "新单价": 18,
    "原材料费": 18,
    "重算材料费": 18,
    "原账单金额": 20,
    "重算后金额": 20,
    "应补金额": 18,
    "计费时间": 22,
    "订单号": 28,
}


def build_bill_workbook(account, summary, detail):
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Bill汇总"
    _write_sheet(summary_sheet, f"{account} Bill 汇总", summary, total_row=True)
    _apply_summary_formulas(summary_sheet, summary)
    detail_sheet = workbook.create_sheet("平台订单明细")
    _write_sheet(detail_sheet, f"{account} 平台订单明细", detail)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_sheet(sheet, title, frame, total_row=False):
    clean = _excel_safe_frame(frame)
    column_count = max(1, len(clean.columns))
    sheet.merge_cells(
        start_row=1, start_column=1, end_row=1, end_column=column_count
    )
    title_cell = sheet.cell(1, 1, title)
    title_cell.font = Font(size=16, bold=True, color="1F1F1F")
    title_cell.alignment = Alignment(horizontal="left")
    header_row = 3
    for column_index, name in enumerate(clean.columns, start=1):
        cell = sheet.cell(header_row, column_index, str(name))
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_index, values in enumerate(
        clean.itertuples(index=False, name=None), start=4
    ):
        for column_index, value in enumerate(values, start=1):
            column_name = clean.columns[column_index - 1]
            if _is_identifier_column(column_name) and value is not None:
                value = str(value)
            cell = sheet.cell(row_index, column_index, value)
            cell.alignment = Alignment(vertical="center")
            if _is_identifier_column(column_name):
                cell.number_format = "@"
            elif str(column_name) == "计费时间":
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif _is_amount_column(column_name):
                cell.number_format = '#,##0.0000;[Red](#,##0.0000);-'
            elif _is_count_column(column_name):
                cell.number_format = '#,##0'
    last_row = header_row + len(clean)
    if total_row and len(clean):
        for cell in sheet[last_row]:
            cell.fill = TOTAL_FILL
            cell.font = Font(bold=True)
    if len(clean):
        sheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(column_count)}{last_row}"
        )
    sheet.freeze_panes = "A4"
    sheet.sheet_view.showGridLines = False
    _set_column_widths(sheet, clean)
    _highlight_amount_due(sheet, clean, total_row=total_row)


def _excel_safe_frame(frame):
    clean = frame.copy()
    if "计费时间" in clean:
        values = clean["计费时间"]
        if pd.api.types.is_numeric_dtype(values):
            values = pd.to_datetime(values, unit="ms", errors="coerce", utc=True)
        else:
            values = pd.to_datetime(values, errors="coerce", utc=True)
        clean["计费时间"] = values.dt.tz_convert(NEW_YORK).dt.tz_localize(None)
    clean = clean.where(pd.notna(clean), None)
    return clean


def _set_column_widths(sheet, frame):
    for index, column in enumerate(frame.columns, start=1):
        values = [
            str(column),
            *(str(value) for value in frame[column].head(200) if value is not None),
        ]
        width = min(max(len(value) for value in values) + 3, 38)
        minimum = MINIMUM_COLUMN_WIDTHS.get(str(column), 10)
        sheet.column_dimensions[get_column_letter(index)].width = max(
            width, minimum
        )


def _apply_summary_formulas(sheet, frame):
    if frame.empty:
        return
    headers = {
        str(sheet.cell(3, column).value): column
        for column in range(1, sheet.max_column + 1)
    }
    required = {
        "商品数量", "正确单价", "原材料费", "原账单金额",
        "重算后金额", "应补金额",
    }
    if not required.issubset(headers):
        return
    total_row = 3 + len(frame)
    data_end = total_row - 1
    quantity = get_column_letter(headers["商品数量"])
    price = get_column_letter(headers["正确单价"])
    original_material = get_column_letter(headers["原材料费"])
    original_total = get_column_letter(headers["原账单金额"])
    recalculated = get_column_letter(headers["重算后金额"])
    amount_due = get_column_letter(headers["应补金额"])
    for row in range(4, total_row):
        price_value = sheet[f"{price}{row}"].value
        if isinstance(price_value, (int, float)):
            sheet[f"{recalculated}{row}"] = (
                f"={original_total}{row}-{original_material}{row}"
                f"+{quantity}{row}*{price}{row}"
            )
            sheet[f"{amount_due}{row}"] = (
                f"={recalculated}{row}-{original_total}{row}"
            )
    if data_end >= 4:
        for name in (
            "商品数量", "原材料费", "原账单金额", "重算后金额", "应补金额",
        ):
            column = get_column_letter(headers[name])
            sheet[f"{column}{total_row}"] = f"=SUM({column}4:{column}{data_end})"


def _highlight_amount_due(sheet, frame, total_row=False):
    if "应补金额" not in frame.columns:
        return
    column = frame.columns.get_loc("应补金额") + 1
    header = sheet.cell(3, column)
    header.fill = AMOUNT_DUE_HEADER_FILL
    header.font = WHITE_BOLD
    last_row = 3 + len(frame)
    for row in range(4, last_row + 1):
        cell = sheet.cell(row, column)
        cell.fill = (
            AMOUNT_DUE_TOTAL_FILL
            if total_row and row == last_row
            else AMOUNT_DUE_FILL
        )
        cell.font = AMOUNT_DUE_FONT


def _is_amount_column(name):
    return any(term in str(name) for term in ("金额", "单价", "材料费"))


def _is_count_column(name):
    return any(term in str(name) for term in ("订单数", "数量"))


def _is_identifier_column(name):
    return str(name) in {"订单号", "客户编号"}
