from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from automation.price_catalogs.haloopod import CATALOG_VERSION, PRICE_CATALOG_ROWS


def build_haloopod_price_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "最高价档"
    sheet.append([CATALOG_VERSION])
    sheet.merge_cells("A1:E1")
    sheet["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    sheet["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    sheet["A1"].alignment = Alignment(horizontal="center")
    sheet.append(["材质", "颜色范围", "尺码", "单面/背面", "双面"])
    for cell in sheet[2]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="5B9BD5")
    for row in PRICE_CATALOG_ROWS:
        sheet.append(list(row))
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:E{sheet.max_row}"
    for column, width in zip("ABCDE", [24, 20, 14, 14, 14]):
        sheet.column_dimensions[column].width = width
    for row in sheet.iter_rows(min_row=3, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = "0.00"
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
