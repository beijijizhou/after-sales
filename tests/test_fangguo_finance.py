import unittest
from io import BytesIO
from unittest.mock import patch

import pandas as pd
from openpyxl import load_workbook

from automation.api.fangguo.finance import (
    PAGE_SIZE,
    _finance_payload,
    _sku_price_payload,
    apply_current_sku_prices,
    build_price_rule_table,
    build_customer_bill_summary,
    build_customer_bill_table,
    normalize_fangguo_finance_lines,
    normalize_fangguo_sku_prices,
    recalculate_fangguo_finance,
    update_fangguo_sku_prices,
)
from automation.api.fangguo.auth import (
    clear_fangguo_login_cache,
    login_fangguo_cached,
)
from automation.api.fangguo.price_catalog import latest_apparel_target_price
from ui.finance.platform_finance import (
    _normalize_price_rules,
    _parse_group_ids,
)
from ui.finance.fangguo_sku_pricing import (
    build_bulk_price_changes,
    build_sku_price_changes,
)
from ui.finance.fangguo_sku_catalog import _upgrade_cached_rows
from ui.finance.bill_workbook import build_bill_workbook


def _api_row(**values):
    return [
        {"label": key, "prop": key, "value": value}
        for key, value in values.items()
    ]


class FangguoFinanceTests(unittest.TestCase):
    def test_old_sku_cache_is_upgraded_with_technology_fields(self):
        result = _upgrade_cached_rows(pd.DataFrame([{"skuId": 1}]))
        self.assertEqual(result.iloc[0]["technologyName"], "")
        self.assertEqual(result.iloc[0]["itemCode"], "")

    def test_latest_catalog_treats_blank_and_back_as_single_only_double_as_double(self):
        base = {"materialCode": "CVC_男T", "colorCode": "黑色", "modelCode": "S"}
        self.assertEqual(latest_apparel_target_price({**base, "technologyName": ""}), 17.5)
        self.assertEqual(latest_apparel_target_price({**base, "technologyName": "背面"}), 17.5)
        self.assertEqual(latest_apparel_target_price({**base, "technologyName": "双面"}), 19.5)

    def test_latest_catalog_uses_exact_price_even_when_it_decreases(self):
        row = {"materialCode": "女连帽卫衣加绒", "colorCode": "黑色", "modelCode": "M", "technologyName": ""}
        self.assertEqual(latest_apparel_target_price(row), 47)

    def test_finance_page_uses_provider_validated_large_page_size(self):
        self.assertEqual(PAGE_SIZE, 2_000)

    def test_normalizes_dynamic_cells_and_deduplicates_order_sku(self):
        first = _api_row(
            tid="ORDER-1", materialCode="杯", colorCode="白",
            modelCode="20oz", skuPropertiesName="304", quantity=1,
            caseAmount=33, totalAmount=33,
        )
        corrected = _api_row(
            tid="ORDER-1", materialCode="杯", colorCode="白",
            modelCode="20oz", skuPropertiesName="304", quantity=1,
            caseAmount=30, totalAmount=30,
        )
        result = normalize_fangguo_finance_lines([first, corrected])
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]["caseAmount"], 30)

    def test_price_rules_ignore_specification_for_same_material_price(self):
        lines = pd.DataFrame([
            {
                "materialCode": "杯", "colorCode": "白",
                "modelCode": "20oz", "skuPropertiesName": "304",
                "quantity": 2, "caseAmount": 66,
            },
            {
                "materialCode": "杯", "colorCode": "白",
                "modelCode": "20oz", "skuPropertiesName": "陶瓷",
                "quantity": 1, "caseAmount": 38,
            },
        ])
        result = build_price_rule_table(lines)
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]["currentUnitPrice"], "33.0000 / 38.0000")

    def test_normalizes_active_fangguo_sku_prices(self):
        result = normalize_fangguo_sku_prices([
            {
                "id": 1, "materialCode": "女收腰短袖",
                "colorCode": "Pink", "modelCode": "S",
                "price": 24, "status": True, "updateTime": 123,
            },
            {
                "id": 2, "materialCode": "女收腰短袖",
                "colorCode": "灰色", "modelCode": "M",
                "price": 99, "status": False, "updateTime": 456,
            },
        ])
        self.assertEqual(len(result), 1)
        self.assertEqual(result.iloc[0]["currentSkuPrice"], 24)

    def test_sku_catalog_can_include_inactive_rows_for_audit(self):
        result = normalize_fangguo_sku_prices([
            {
                "id": 1, "materialCode": "男T", "colorCode": "黑",
                "modelCode": "S", "price": 10, "status": True,
            },
            {
                "id": 2, "materialCode": "男T", "colorCode": "黑",
                "modelCode": "M", "price": 10, "status": False,
            },
        ], include_inactive=True)
        self.assertEqual(len(result), 2)
        self.assertEqual(int(result["skuActive"].sum()), 1)

    def test_sku_catalog_preserves_same_identity_with_different_print_faces(self):
        result = normalize_fangguo_sku_prices([
            {"id": 1, "materialCode": "男T", "colorCode": "黑", "modelCode": "S", "technologyName": "背面", "price": 17.5, "status": True},
            {"id": 2, "materialCode": "男T", "colorCode": "黑", "modelCode": "S", "technologyName": "双面", "price": 19.5, "status": True},
        ], include_inactive=True)
        self.assertEqual(len(result), 2)
        self.assertEqual(result["technologyName"].tolist(), ["背面", "双面"])

    def test_builds_only_selected_changed_sku_prices(self):
        source = normalize_fangguo_sku_prices([{
            "id": 12101698, "materialCode": "磨砂TPU",
            "colorCode": "黑色", "modelCode": "iphone12promax",
            "price": 11, "status": True,
        }], include_inactive=True)
        editor = pd.DataFrame([{
            "选择": True, "方果 SKU ID": 12101698,
            "材质 / 商品": "磨砂TPU", "颜色": "黑色",
            "型号 / 尺码": "iphone12promax",
            "当前价格": 11, "新价格": 12,
        }])
        changes = build_sku_price_changes(source, editor)
        self.assertEqual(len(changes), 1)
        self.assertEqual(changes.iloc[0]["newPrice"], 12)
        self.assertEqual(changes.iloc[0]["sourcePayload"]["id"], 12101698)

    def test_bulk_fixed_increase_applies_without_per_row_input(self):
        source = normalize_fangguo_sku_prices([
            {"id": 1, "materialCode": "男T", "colorCode": "黑", "modelCode": "S", "price": 10, "status": True},
            {"id": 2, "materialCode": "男T", "colorCode": "黑", "modelCode": "M", "price": 12, "status": True},
        ], include_inactive=True)
        changes = build_bulk_price_changes(source, [1, 2], "fixed", 1.5)
        self.assertEqual(changes["newPrice"].tolist(), [11.5, 13.5])
        self.assertEqual(changes["increase"].tolist(), [1.5, 1.5])

    def test_bulk_percentage_increase_uses_each_current_price(self):
        source = normalize_fangguo_sku_prices([
            {"id": 1, "materialCode": "男T", "colorCode": "黑", "modelCode": "S", "price": 10, "status": True},
            {"id": 2, "materialCode": "男T", "colorCode": "黑", "modelCode": "M", "price": 20, "status": True},
        ], include_inactive=True)
        changes = build_bulk_price_changes(source, [1, 2], "percent", 10)
        self.assertEqual(changes["newPrice"].tolist(), [11, 22])
        self.assertEqual(changes["increase"].tolist(), [1, 2])

    def test_old_cached_sku_rows_without_source_payload_do_not_crash(self):
        old_cache = pd.DataFrame([{
            "skuId": 1, "materialCode": "男T", "colorCode": "黑",
            "modelCode": "S", "currentSkuPrice": 10,
        }])
        changes = build_bulk_price_changes(old_cache, [1], "fixed", 1)
        self.assertTrue(changes.empty)

    def test_update_sku_price_preserves_source_payload_and_changes_price(self):
        class Response:
            def raise_for_status(self):
                return None

            def json(self):
                return {"code": 0, "msg": ""}

        class Client:
            def __init__(self):
                self.calls = []

            def post(self, url, **kwargs):
                self.calls.append((url, kwargs))
                return Response()

        client = Client()
        credentials = {"tenant_id": "tenant", "token": "token"}
        source = {
            "id": 12101698, "materialCode": "磨砂TPU",
            "modelCode": "iphone12promax", "price": 11,
        }
        with patch(
            "automation.api.fangguo.finance._authenticated_client",
            return_value=(client, "token"),
        ):
            result = update_fangguo_sku_prices(credentials, [{
                "skuId": 12101698, "newPrice": 12,
                "sourcePayload": source,
            }])
        self.assertTrue(result.iloc[0]["success"])
        sent = client.calls[0][1]["json"]
        self.assertEqual(sent["materialCode"], "磨砂TPU")
        self.assertEqual(sent["skuId"], 12101698)
        self.assertEqual(sent["price"], 12)

    def test_current_sku_price_prefills_recalculation_price(self):
        lines = pd.DataFrame([{
            "materialCode": "女收腰短袖", "quantity": 1,
            "caseAmount": 20,
        }])
        sku_prices = normalize_fangguo_sku_prices([
            {
                "id": 1, "materialCode": "女收腰短袖",
                "colorCode": "Pink", "modelCode": "S",
                "price": 24, "status": True,
            },
            {
                "id": 2, "materialCode": "女收腰短袖",
                "colorCode": "灰色", "modelCode": "M",
                "price": 24, "status": True,
            },
        ])
        rules = apply_current_sku_prices(
            build_price_rule_table(lines), sku_prices
        )
        self.assertEqual(rules.iloc[0]["fangguoSkuPrice"], "24.0000")
        self.assertEqual(rules.iloc[0]["newUnitPrice"], 24)

    def test_conflicting_current_sku_prices_require_manual_review(self):
        lines = pd.DataFrame([{
            "materialCode": "女收腰短袖", "quantity": 1,
            "caseAmount": 20,
        }])
        sku_prices = normalize_fangguo_sku_prices([
            {
                "id": 1, "materialCode": "女收腰短袖",
                "colorCode": "Pink", "modelCode": "S",
                "price": 24, "status": True,
            },
            {
                "id": 2, "materialCode": "女收腰短袖",
                "colorCode": "灰色", "modelCode": "M",
                "price": 26, "status": True,
            },
        ])
        rules = apply_current_sku_prices(
            build_price_rule_table(lines), sku_prices
        )
        self.assertEqual(rules.iloc[0]["fangguoSkuPrice"], "24.0000 / 26.0000")
        self.assertTrue(pd.isna(rules.iloc[0]["newUnitPrice"]))

    def test_recalculation_preserves_non_material_fees(self):
        lines = pd.DataFrame([{
            "tid": "ORDER-1", "materialCode": "杯", "colorCode": "白",
            "modelCode": "20oz", "skuPropertiesName": "304",
            "quantity": 2, "caseAmount": 66, "shippingAmount": 5,
            "totalAmount": 71,
        }])
        rules = pd.DataFrame([{
            "materialCode": "杯", "colorCode": "白",
            "modelCode": "20oz", "skuPropertiesName": "304",
            "newUnitPrice": 30,
        }])
        result = recalculate_fangguo_finance(lines, rules).iloc[0]
        self.assertEqual(result["recalculatedCaseAmount"], 60)
        self.assertEqual(result["difference"], -6)
        self.assertEqual(result["recalculatedTotalAmount"], 65)

    def test_one_material_price_applies_across_specs(self):
        lines = pd.DataFrame([
            {
                "tid": f"ORDER-{spec}", "materialCode": "男士T恤",
                "colorCode": color, "modelCode": "", "skuPropertiesName": spec,
                "quantity": 1, "caseAmount": old, "totalAmount": old,
            }
            for spec, color, old in [("S", "黑", 8), ("XL", "白", 9)]
        ])
        rules = pd.DataFrame([{"materialCode": "男士T恤", "newUnitPrice": 10}])
        result = recalculate_fangguo_finance(lines, rules)
        self.assertEqual(result["recalculatedCaseAmount"].tolist(), [10, 10])
        self.assertEqual(result["difference"].tolist(), [2, 1])

    def test_old_rule_editor_columns_are_removed(self):
        lines = pd.DataFrame([{
            "materialCode": "男士T恤", "quantity": 1, "caseAmount": 8,
        }])
        old = pd.DataFrame([{
            "materialCode": "男士T恤", "colorCode": "黑",
            "modelCode": "S", "skuPropertiesName": "旧规格",
            "newUnitPrice": 10,
        }])
        normalized = _normalize_price_rules(lines, old, ["materialCode"])
        self.assertEqual(
            normalized.columns.tolist(),
            ["materialCode", "currentUnitPrice", "newUnitPrice"],
        )
        self.assertEqual(normalized.iloc[0]["newUnitPrice"], 10)

    def test_user_can_choose_model_specific_pricing(self):
        lines = pd.DataFrame([
            {
                "materialCode": "硅藻泥地垫", "modelCode": "40x60",
                "quantity": 1, "caseAmount": 8, "totalAmount": 8,
            },
            {
                "materialCode": "硅藻泥地垫", "modelCode": "50x80",
                "quantity": 1, "caseAmount": 9, "totalAmount": 9,
            },
        ])
        rules = build_price_rule_table(lines, ["materialCode", "modelCode"])
        self.assertEqual(len(rules), 2)
        rules.loc[rules["modelCode"] == "40x60", "newUnitPrice"] = 10
        rules.loc[rules["modelCode"] == "50x80", "newUnitPrice"] = 12
        result = recalculate_fangguo_finance(
            lines, rules, ["materialCode", "modelCode"]
        )
        self.assertEqual(result["recalculatedCaseAmount"].tolist(), [10, 12])

    def test_customer_bills_combine_longfeng_and_separate_haloo(self):
        result = pd.DataFrame([
            {
                "shopName": "隆丰1", "shopCode": "隆丰1", "tid": "L1",
                "quantity": 2, "totalAmount": 16,
                "recalculatedTotalAmount": 20, "difference": 4,
            },
            {
                "shopName": "隆丰3", "shopCode": "", "tid": "L3",
                "quantity": 1, "totalAmount": 8,
                "recalculatedTotalAmount": 10, "difference": 2,
            },
            {
                "shopName": "海捞", "shopCode": "haloo", "tid": "H1",
                "quantity": 3, "totalAmount": 24,
                "recalculatedTotalAmount": 30, "difference": 6,
            },
        ])
        bills = build_customer_bill_summary(result).set_index("customerAccount")
        self.assertEqual(bills.loc["隆丰", "orderCount"], 2)
        self.assertEqual(bills.loc["隆丰", "amountDue"], 6)
        self.assertEqual(bills.loc["Haloo", "amountDue"], 6)

    def test_customer_bill_table_is_product_summary_with_total(self):
        result = pd.DataFrame([
            {
                "shopName": "海捞", "shopCode": "haloo", "tid": "H1",
                "materialCode": "男士T恤", "modelCode": "S",
                "quantity": 2, "newUnitPrice": 10,
                "caseAmount": 16, "totalAmount": 16,
                "recalculatedTotalAmount": 20,
                "difference": 4,
            },
            {
                "shopName": "海捞", "shopCode": "haloo", "tid": "H2",
                "materialCode": "男士T恤", "modelCode": "M",
                "quantity": 1, "newUnitPrice": 10,
                "caseAmount": 8, "totalAmount": 8,
                "recalculatedTotalAmount": 10,
                "difference": 2,
            },
        ])
        bill = build_customer_bill_table(result, "Haloo")
        self.assertEqual(bill["materialCode"].tolist(), ["男士T恤", "男士T恤", "合计"])
        self.assertEqual(bill.iloc[-1]["orderCount"], 2)
        self.assertEqual(bill.iloc[-1]["quantity"], 3)
        self.assertEqual(bill.iloc[-1]["amountDue"], 6)

        combined = build_customer_bill_table(
            result, "Haloo", include_model=False
        )
        self.assertEqual(combined["materialCode"].tolist(), ["男士T恤", "合计"])
        self.assertEqual(combined.iloc[0]["quantity"], 3)

    def test_bill_workbook_has_summary_and_order_detail_sheets(self):
        summary = pd.DataFrame([
            {
                "商品 / 材质": "男士T恤", "订单数": 2, "商品数量": 3,
                "正确单价": 10, "原材料费": 24, "原账单金额": 24,
                "重算后金额": 30, "应补金额": 6,
            },
            {
                "商品 / 材质": "合计", "订单数": 2, "商品数量": 3,
                "正确单价": "", "原材料费": 24, "原账单金额": 24,
                "重算后金额": 30, "应补金额": 6,
            },
        ])
        detail = pd.DataFrame([{
            "计费时间": 1786716000000, "客户": "海捞",
            "订单号": "000577522067710185781", "商品 / 材质": "男士T恤", "数量": 3,
            "原材料费": 24, "新单价": 10, "重算材料费": 30,
            "原账单金额": 24, "应补金额": 6, "重算后金额": 30,
        }])
        content = build_bill_workbook("Haloo", summary, detail)
        workbook = load_workbook(BytesIO(content), data_only=False)
        self.assertEqual(workbook.sheetnames, ["Bill汇总", "平台订单明细"])
        self.assertEqual(workbook["Bill汇总"]["A1"].value, "Haloo Bill 汇总")
        self.assertEqual(workbook["Bill汇总"]["G4"].value, "=F4-E4+C4*D4")
        self.assertEqual(workbook["Bill汇总"]["H4"].value, "=G4-F4")
        self.assertEqual(workbook["Bill汇总"]["G5"].value, "=SUM(G4:G4)")
        order_cell = workbook["平台订单明细"]["C4"]
        self.assertEqual(order_cell.value, "000577522067710185781")
        self.assertEqual(order_cell.data_type, "s")
        self.assertEqual(order_cell.number_format, "@")
        time_cell = workbook["平台订单明细"]["A4"]
        self.assertEqual(time_cell.value.strftime("%Y-%m-%d %H:%M:%S"), "2026-08-14 10:00:00")
        self.assertEqual(time_cell.number_format, "yyyy-mm-dd hh:mm:ss")
        self.assertGreaterEqual(
            workbook["Bill汇总"].column_dimensions["E"].width, 18
        )
        summary_due = workbook["Bill汇总"]
        self.assertEqual(summary_due["H3"].fill.fgColor.rgb, "00C00000")
        self.assertEqual(summary_due["H4"].fill.fgColor.rgb, "00FCE8E6")
        self.assertEqual(summary_due["H4"].font.color.rgb, "00C00000")
        self.assertTrue(summary_due["H4"].font.bold)
        self.assertEqual(summary_due["H5"].fill.fgColor.rgb, "00F4CCCC")
        detail_due = workbook["平台订单明细"]["J4"]
        self.assertEqual(detail_due.fill.fgColor.rgb, "00FCE8E6")
        self.assertEqual(detail_due.font.color.rgb, "00C00000")

    def test_blank_new_price_keeps_platform_amount(self):
        lines = pd.DataFrame([{
            "materialCode": "杯", "colorCode": "白", "modelCode": "20oz",
            "skuPropertiesName": "304", "quantity": 2,
            "caseAmount": 66, "totalAmount": 71,
        }])
        rules = build_price_rule_table(lines)
        result = recalculate_fangguo_finance(lines, rules).iloc[0]
        self.assertEqual(result["recalculatedTotalAmount"], 71)
        self.assertFalse(result["priceRuleApplied"])

    def test_payload_keeps_customer_group_and_all_finance_fee_types(self):
        payload = _finance_payload(1, 500, 100, 200, [8564])
        self.assertEqual(payload["groupIdList"], [8564])
        self.assertIn("ACTUAL_DEDUCTION", payload["feeTypeList"])
        self.assertIn("REFUND", payload["feeTypeList"])

    def test_sku_price_payload_uses_configured_scope(self):
        payload = _sku_price_payload(1, 2000, 36, [177714, 177715], [32494])
        self.assertEqual(payload["materialIds"], [177714, 177715])
        self.assertEqual(payload["colorIds"], [32494])
        self.assertEqual(payload["sortField"], "updateTime")

    def test_blank_customer_group_queries_all_customers(self):
        self.assertEqual(_parse_group_ids(""), [])
        payload = _finance_payload(1, 500, 100, 200, [])
        self.assertEqual(payload["groupIdList"], [])

    def test_multiple_customer_groups_are_supported(self):
        self.assertEqual(_parse_group_ids("8564, 7721"), [8564, 7721])

    def test_customer_groups_accept_streamlit_stringified_list(self):
        self.assertEqual(_parse_group_ids("[8564, 7721]"), [8564, 7721])

    def test_customer_groups_accept_toml_list(self):
        self.assertEqual(_parse_group_ids([8564, "7721"]), [8564, 7721])

    def test_login_token_is_reused_during_cache_window(self):
        credentials = {
            "tenant_id": "tenant", "username": "user",
            "password": "secret", "token_cache_seconds": 2700,
        }
        clear_fangguo_login_cache(credentials)
        with patch(
            "automation.api.fangguo.auth.login_fangguo",
            return_value=(object(), "cached-token"),
        ) as login:
            first = login_fangguo_cached(credentials)
            second = login_fangguo_cached(credentials)
        self.assertEqual(first[1], "cached-token")
        self.assertEqual(second[1], "cached-token")
        login.assert_called_once_with(credentials)
        clear_fangguo_login_cache(credentials)


if __name__ == "__main__":
    unittest.main()
